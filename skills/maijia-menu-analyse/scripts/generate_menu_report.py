#!/usr/bin/env python3
"""Generate a Maijia-style menu analysis report from Meituan cost/gross-profit export."""

from __future__ import annotations

import argparse
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


PRICE_BINS = [
    ("未定价", None, 0),
    ("<10元", 0, 10),
    ("10-20元", 10, 20),
    ("20-30元", 20, 30),
    ("30-40元", 30, 40),
    ("40-50元", 40, 50),
    ("50-60元", 50, 60),
    ("60-80元", 60, 80),
    ("80-100元", 80, 100),
    ("100-150元", 100, 150),
    (">=150元", 150, None),
]

MARGIN_BINS = [
    ("无毛利率", None, None),
    ("<50%", None, 0.5),
    ("50-60%", 0.5, 0.6),
    ("60-70%", 0.6, 0.7),
    ("70-80%", 0.7, 0.8),
    ("80-85%", 0.8, 0.85),
    ("85-90%", 0.85, 0.9),
    ("90-95%", 0.9, 0.95),
    (">=95%", 0.95, None),
]

REQUIRED_COLUMNS = [
    "品项名称",
    "品项类型",
    "菜品编码",
    "单位",
    "销售数量",
    "销售收入(元)",
    "折后均价(元)",
    "菜品预估成本(元)",
    "折后毛利(元)",
    "折后毛利率",
    "优惠金额(元)",
]

TITLE_FILL = "7A4A16"
DARK_FILL = "3F2F1C"
TAN_FILL = "F6E3BD"
LIGHT_FILL = "FFF4CC"
BORDER = Side(style="thin", color="D9C6A5")


def to_number(value: Any) -> float:
    if value in (None, "", "--"):
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def price_bucket(price: float) -> str:
    if price == 0:
        return "未定价"
    for label, min_value, max_value in PRICE_BINS:
        if min_value is not None and price >= min_value and (max_value is None or price < max_value):
            return label
    return "未知"


def margin_bucket(value: Any) -> str:
    if value in (None, "", "--"):
        return "无毛利率"
    margin = to_number(value)
    for label, min_value, max_value in MARGIN_BINS[1:]:
        if (min_value is None or margin >= min_value) and (max_value is None or margin < max_value):
            return label
    return "未知"


def read_source(input_path: Path) -> tuple[str, list[dict[str, Any]]]:
    wb = load_workbook(input_path, data_only=True, read_only=False)
    ws = wb.active
    max_col = max(ws.max_column, 25)
    meta = str(ws.cell(2, 1).value or "")
    headers = [str(ws.cell(3, col).value or "").strip() for col in range(1, max_col + 1)]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise SystemExit(f"Missing required columns: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for row_idx in range(4, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]
        record = dict(zip(headers, values))
        if not record.get("品项名称") or record.get("品项名称") == "合计":
            continue
        rows.append(record)
    return meta, rows


def enrich(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    enriched = []
    for index, row in enumerate(rows, start=1):
        revenue = to_number(row["销售收入(元)"])
        gross_profit = to_number(row["折后毛利(元)"])
        item = {
            "序号": index,
            "品项名称": row["品项名称"],
            "品项类型": row["品项类型"],
            "菜品编码": row["菜品编码"],
            "单位": row["单位"],
            "销售数量": to_number(row["销售数量"]),
            "销售收入": revenue,
            "折后均价": to_number(row["折后均价(元)"]),
            "菜品预估成本": to_number(row["菜品预估成本(元)"]),
            "折后毛利": gross_profit,
            "折后毛利率": gross_profit / revenue if revenue else None,
            "优惠金额": to_number(row["优惠金额(元)"]),
            "价格区间": price_bucket(to_number(row["折后均价(元)"])),
            "毛利区间": margin_bucket(row["折后毛利率"]),
        }
        enriched.append(item)
    return enriched


def aggregate(rows: list[dict[str, Any]], labels: list[str], key: str, total_revenue: float) -> list[dict[str, Any]]:
    grouped = OrderedDict(
        (label, {"分组": label, "销售收入": 0.0, "销售数量": 0.0, "菜品预估成本": 0.0, "折后毛利": 0.0, "品项数": 0})
        for label in labels
    )
    for row in rows:
        label = row[key]
        grouped.setdefault(label, {"分组": label, "销售收入": 0.0, "销售数量": 0.0, "菜品预估成本": 0.0, "折后毛利": 0.0, "品项数": 0})
        bucket = grouped[label]
        bucket["销售收入"] += row["销售收入"]
        bucket["销售数量"] += row["销售数量"]
        bucket["菜品预估成本"] += row["菜品预估成本"]
        bucket["折后毛利"] += row["折后毛利"]
        bucket["品项数"] += 1
    output = []
    for row in grouped.values():
        if row["品项数"] == 0:
            continue
        row["折后毛利率"] = row["折后毛利"] / row["销售收入"] if row["销售收入"] else None
        row["销售占比"] = row["销售收入"] / total_revenue if total_revenue else 0
        output.append(row)
    return output


def set_title(ws, title: str, subtitle: str, end_col: str) -> None:
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A2:{end_col}2")
    ws["A2"] = subtitle
    ws["A2"].fill = PatternFill("solid", fgColor=TAN_FILL)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")


def style_header(row) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=DARK_FILL)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")


def style_range(ws, min_row: int, max_row: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
            cell.alignment = Alignment(vertical="center")


def write_analysis_sheet(ws, title: str, rows: list[dict[str, Any]], note: str) -> None:
    set_title(ws, title, "可用维度 = 价格区间 + 毛利区间", "H")
    headers = ["分组", "销售收入", "销售数量", "菜品预估成本", "折后毛利", "折后毛利率", "销售占比", "品项数"]
    ws.append(headers)
    style_header(ws[3])
    for row in rows:
        ws.append([row[h] for h in headers])
    note_row = ws.max_row + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
    ws.cell(note_row, 1).value = note
    ws.cell(note_row, 1).font = Font(italic=True, color="475569")
    ws.freeze_panes = "A4"
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 15
    for row in ws.iter_rows(min_row=4, max_row=note_row - 2, min_col=2, max_col=5):
        for cell in row:
            cell.number_format = "#,##0.00"
    for row in ws.iter_rows(min_row=4, max_row=note_row - 2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "0.0%"
    style_range(ws, 3, note_row - 2, 8)


def build_report(input_path: Path, output_path: Path, title: str) -> None:
    meta, raw_rows = read_source(input_path)
    rows = enrich(raw_rows)
    total_revenue = sum(row["销售收入"] for row in rows)
    total_cost = sum(row["菜品预估成本"] for row in rows)
    total_profit = sum(row["折后毛利"] for row in rows)
    total_qty = sum(row["销售数量"] for row in rows)
    total_discount = sum(row["优惠金额"] for row in rows)
    price_rows = aggregate(rows, [label for label, _, _ in PRICE_BINS], "价格区间", total_revenue)
    margin_rows = aggregate(rows, [label for label, _, _ in MARGIN_BINS], "毛利区间", total_revenue)

    wb = Workbook()
    wb.remove(wb.active)
    summary = wb.create_sheet("2A_总览")
    price_sheet = wb.create_sheet("A1_价格区间分析")
    margin_sheet = wb.create_sheet("A2_毛利区间分析")
    detail = wb.create_sheet("完整菜品清单")

    set_title(summary, title, "现有字段分析 = 价格区间 + 毛利区间", "H")
    summary.append([])
    summary.append(["销售收入", "菜品预估成本", "折后毛利", "折后毛利率", "品项数", "销售数量", "优惠金额", "数据口径"])
    summary.append([total_revenue, total_cost, total_profit, total_profit / total_revenue if total_revenue else None, len(rows), total_qty, total_discount, "按导出文件"])
    style_header(summary[4])
    for cell in summary[5][:3] + (summary[5][6],):
        cell.number_format = "#,##0"
    summary["D5"].number_format = "0.0%"
    summary["F5"].number_format = "#,##0"
    summary.append([])
    summary.append(["维度", "分组", "销售收入", "销售数量", "折后毛利", "折后毛利率", "销售占比", "品项数"])
    style_header(summary[7])
    for row in price_rows:
        summary.append(["A1 价格区间", row["分组"], row["销售收入"], row["销售数量"], row["折后毛利"], row["折后毛利率"], row["销售占比"], row["品项数"]])
    summary.append([None] * 8)
    for row in margin_rows:
        summary.append(["A2 毛利区间", row["分组"], row["销售收入"], row["销售数量"], row["折后毛利"], row["折后毛利率"], row["销售占比"], row["品项数"]])
    for row in summary.iter_rows(min_row=8, max_row=summary.max_row, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = "#,##0"
    for row in summary.iter_rows(min_row=8, max_row=summary.max_row, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "0.0%"
    for col in "ABCDEFGH":
        summary.column_dimensions[col].width = 15
    summary.freeze_panes = "A8"

    write_analysis_sheet(price_sheet, "A1_价格区间分析", price_rows, "说明：成本毛利表没有菜单标价，本页价格区间按“折后均价(元)”推导。")
    write_analysis_sheet(margin_sheet, "A2_毛利区间分析", margin_rows, "说明：本页毛利区间按“折后毛利率”推导；无销售收入或系统未给出毛利率的品项归为“无毛利率”。")

    set_title(detail, "完整菜品清单（基于当前成本毛利表）", f"来源口径：{meta[:120]}", "L")
    detail_headers = ["序号", "品项名称", "价格区间", "毛利区间", "销售数量", "销售收入", "菜品预估成本", "折后毛利", "折后毛利率", "销售占比", "品项类型", "单位"]
    detail.append(detail_headers)
    style_header(detail[3])
    for idx, row in enumerate(sorted(rows, key=lambda item: item["销售收入"], reverse=True), start=1):
        detail.append([
            idx,
            row["品项名称"],
            row["价格区间"],
            row["毛利区间"],
            row["销售数量"],
            row["销售收入"],
            row["菜品预估成本"],
            row["折后毛利"],
            row["折后毛利率"],
            row["销售收入"] / total_revenue if total_revenue else 0,
            row["品项类型"],
            row["单位"],
        ])
    for row in detail.iter_rows(min_row=4, max_row=detail.max_row, min_col=5, max_col=8):
        for cell in row:
            cell.number_format = "#,##0.00"
    for row in detail.iter_rows(min_row=4, max_row=detail.max_row, min_col=9, max_col=10):
        for cell in row:
            cell.number_format = "0.0%"
    detail.column_dimensions["B"].width = 24
    for col in "ACDEFGHIJKL":
        detail.column_dimensions[col].width = 14
    detail.freeze_panes = "A4"
    table_ref = f"A3:L{detail.max_row}"
    table = Table(displayName="MenuDetail", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    detail.add_table(table)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--title", default="永杰厚道 菜单分析报告")
    args = parser.parse_args()
    build_report(args.input, args.output, args.title)
    print(args.output)


if __name__ == "__main__":
    main()
